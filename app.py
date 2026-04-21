"""
Copyright (c) 2026 건원건축(Kunwon Architecture) & 김정현. All rights reserved.

본 프로그램은 건원건축의 도면 검토 업무 효율화를 위해 기획 및 개발되었습니다.
사내 임직원 외 외부 업체로의 유출, 무단 복제 및 소스코드 수정을 엄격히 금지합니다.

app.py  —  DWG 자동 검토기 v_6.8 (Kunwon Masterpiece - Multi-Zoning & Cleanup)
========================================================================
[V6.8 주요 업데이트]
1. 다단(Multi-Column) ROI 완벽 지원: 리습에서 지정한 N개의 목록표 단(박스)을 
   각각 독립된 구역으로 인식하여 분석합니다.
2. 스마트 정제 엔진 탑재: 
   - 도연번호(오타) 인식
   - 일련번호(도면번호 좌측 텍스트) 절단
   - 비고(Remarks) 차단막 설치 (도면명과 분리)
   - 축척(1:) 등 범례 찌꺼기 완벽 청소
   - 카테고리 행(A0 [공통사항] 등) 자동 스킵
========================================================================
"""

from __future__ import annotations
import glob, os, re, sys, webbrowser, json, math
import tkinter as tk
from tkinter import messagebox
from pathlib import Path
from typing import List, Optional, Tuple
import concurrent.futures

import pandas as pd
import ezdxf
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 리포트 파일 기본 이름
리포트_이름: str = "도면검토리포트_최종.xlsx"
ODA_DOWNLOAD_URL = "https://www.opendesign.com/guestfiles/oda_file_converter"

# ============================================================================
# 0. 설정 로드 및 엔진 세팅
# ============================================================================
def load_roi_config(block_name: str) -> Optional[dict]:
    config_dir = os.path.join(os.environ.get('APPDATA', ''), 'AutoDWG_Checker')
    config_path = os.path.join(config_dir, f"{block_name}.json")
    if os.path.exists(config_path):
        for enc in ['cp949', 'utf-8', 'euc-kr']:
            try:
                with open(config_path, 'r', encoding=enc) as f:
                    return json.load(f)
            except Exception:
                continue
    return None

def _oda_환경_설정() -> str:
    found_path = ""
    for 경로 in [r"C:\Program Files\ODA", r"C:\Program Files (x86)\ODA"]:
        실행파일들 = glob.glob(os.path.join(경로, "**", "ODAFileConverter.exe"), recursive=True)
        if 실행파일들:
            found_path = sorted(실행파일들, reverse=True)[0]
            break
    if found_path:
        폴더경로 = os.path.dirname(found_path)
        if 폴더경로 not in os.environ.get("PATH", ""):
            os.environ["PATH"] = 폴더경로 + os.pathsep + os.environ.get("PATH", "")
        try:
            ezdxf.options.odafc_win_exec_path = found_path
        except AttributeError:
            if not ezdxf.options.has_section('odafc'):
                ezdxf.options._config.add_section('odafc')
            ezdxf.options.set('odafc', 'win_exec_path', found_path)
    return found_path

def check_oda_installation():
    if not _oda_환경_설정():
        root = tk.Tk(); root.withdraw()
        msg = (
            "⚠️ CAD 분석 엔진(ODA)이 설치되어 있지 않습니다!\n\n"
            "확인을 누르면 다운로드 페이지가 열립니다. 기본 경로에 설치해 주세요."
        )
        if messagebox.askokcancel("엔진 설치 안내", msg):
            webbrowser.open(ODA_DOWNLOAD_URL)
        sys.exit()

# ============================================================================
# 1. 공통 유틸리티 (지우개 및 필터)
# ============================================================================
_도면번호_패턴 = re.compile(r"(?<![가-힣A-Za-z0-9])([A-Z\u0391-\u03A9\.가-힣][A-Z0-9\u0391-\u03A9\.가-힣]{0,4})[\s\-_~–—−]*(\d{1,5}(?:[-.~–—−]\d{1,3})?[A-Z]*|TOE)(?!\d|[A-Za-z])")
_축척_패턴 = re.compile(r"(1\s?[/:,]\s?(\d{1,4})|NONE|N/A)", re.I)
_동_패턴 = re.compile(r"((?:(?:[0-9A-Za-z]+|[가-힣]|[0-9A-Za-z가-힣]+동)\s*[,~&]\s*)*[0-9A-Za-z가-힣]+동)")
_동_제외단어 = ["인동", "주동", "공동", "자동", "수동", "전동", "연동", "이동", "작동", "부동", "진동", "명동", "구동", "개동", "각동", "해당동", "상동", "하동"]

GLOBAL_IGNORE_HEADERS = [
    "DRAWING TITLE", "DRAWING NO.", "DRAWING NO", "DWG.NO.", "DWG. NO.", "DWG.NO", "DWG NO.", "DWG NO", "TITLE",
    "도면번호", "도연번호", "일련번호", "연번", "NO", "NO.", "도면명", "도면명칭", "축척(A1)", "축척(A3)", "축척(A0)", 
    "SCALE(A1)", "SCALE(A3)", "SCALE(A0)", "축척(1:)", "축척(1/)", "SCALE(1:)", "SCALE(1/)", "(1:)", "(1/)",
    "축척", "축적", "SCALE", "비고", "REMARK", "REMARKS", "사업승인", "착공", "견적", "사용승인", "1:1", "도면"
]

CATEGORY_KEYWORDS = [
    "공통사항", "일반사항", "건축도면", "구조도면", "기계도면", 
    "전기도면", "토목도면", "조경도면", "소방도면", "부분상세도"
]

def _clean_text_from_headers(txt: str) -> str:
    clean = txt
    for h in sorted(GLOBAL_IGNORE_HEADERS, key=len, reverse=True):
        pattern = re.compile(re.escape(h), re.IGNORECASE)
        clean = pattern.sub(" ", clean)
    clean = re.sub(r"\s+", " ", clean).strip()
    clean = re.sub(r"^[-_,\s]+|[-_,\s]+$", "", clean)
    return clean

def _extract_dong_from_title(title: str) -> str:
    matches = list(_동_패턴.finditer(title))
    for m in matches:
        dong_str = m.group(1)
        is_excluded = False
        for ex_word in _동_제외단어:
            if ex_word in dong_str:
                is_excluded = True
                break
        if not is_excluded:
            return dong_str
    return ""

def _도면번호_세척(raw_s: str) -> str:
    if not raw_s: return ""
    s = raw_s.strip().upper().replace("Λ", "A").replace("Δ", "A").replace("TOE", "108")
    if s.startswith("."): s = "AA" + s[1:]
    return re.sub(r"\s+", " ", s)

def _축척_텍스트_정리(txt: str) -> str:
    if not txt: return "X"
    u = txt.upper()
    if "NONE" in u or "N/A" in u: return "NONE"
    m = _축척_패턴.search(u)
    return f"1/{m.group(2)}" if m and m.group(2) else "X"

def _extract_drawing_number(text: str) -> Optional[str]:
    for m in _도면번호_패턴.finditer(text):
        prefix = m.group(1)
        if m.group(0) in ["A1", "A3", "A0", "A2", "A4"]: continue
        exclude_words = ["상세", "일람", "배치", "전개", "마감", "계획", "조감", "구조", "코어", "지하", "옥상", "옥탑", "지붕", "주동", "단위", "세대", "내역", "관계", "형별", "부분", "창호", "가구", "조경", "토목", "기계", "전기", "범례", "개요", "표지", "도면", "시설", "센터", "주차장", "휴게소", "사무소", "경로당", "어린이집", "유치원", "도서관", "커뮤니티", "피트니스", "사우나", "골프", "문주", "경비실"]
        if any(k in prefix for k in exclude_words): continue
        if prefix.endswith("도") or prefix.endswith("표") or prefix.endswith("층") or prefix.endswith("동"): continue
        return m.group(0)
    return None

def _정리문자열(txt: str) -> str:
    return re.sub(r"\s+", " ", (txt or "")).strip()

def _cad_로드(path: Path):
    if path.suffix.lower() == ".dxf": return ezdxf.readfile(str(path))
    _oda_환경_설정()
    from ezdxf.addons import odafc
    return odafc.readfile(str(path))

def _get_safe_point(ent) -> Tuple[float, float]:
    p = ent.dxf.insert
    if getattr(ent.dxf, "halign", 0) > 0 or getattr(ent.dxf, "valign", 0) > 0:
        ap = getattr(ent.dxf, "align_point", None)
        if ap and (round(ap[0], 2) != 0 or round(ap[1], 2) != 0): p = ap
    return float(p[0]), float(p[1])

def _텍스트_데이터_추출(ent) -> List[Tuple[float, float, str, float]]:
    유형 = ent.dxftype()
    결과 = []
    try:
        if 유형 == "TEXT":
            px, py = _get_safe_point(ent)
            h = getattr(ent.dxf, "height", 10.0)
            txt = (ent.dxf.text or "").strip()
            if txt: 결과.append((px, py, txt, float(h)))
        elif 유형 == "MTEXT":
            h = getattr(ent.dxf, "char_height", 10.0)
            bx, by = float(ent.dxf.insert[0]), float(ent.dxf.insert[1])
            lines = ent.plain_text().split('\n')
            for i, line in enumerate(lines):
                txt = line.strip()
                if txt: 결과.append((bx, by - (i * h * 1.5), txt, float(h)))
        elif 유형 == "ATTRIB":
            px, py = _get_safe_point(ent)
            h = getattr(ent.dxf, "height", 10.0)
            txt = (ent.dxf.text or "").strip()
            if txt: 결과.append((px, py, txt, float(h)))
    except Exception: pass
    return 결과

def _collect_layout_texts(layout) -> List[Tuple[float, float, str, float]]:
    texts = []
    try:
        for ent in layout.query("TEXT MTEXT LINE LWPOLYLINE INSERT"):
            if ent.dxftype() in ["TEXT", "MTEXT", "LINE", "LWPOLYLINE"]:
                texts.extend(_텍스트_데이터_추출(ent))
            elif ent.dxftype() == "INSERT":
                for att in getattr(ent, "attribs", []): texts.extend(_텍스트_데이터_추출(att))
                try:
                    for v_ent in ent.virtual_entities():
                        if v_ent.dxftype() in ["TEXT", "MTEXT", "LINE", "LWPOLYLINE"]: texts.extend(_텍스트_데이터_추출(v_ent))
                        elif v_ent.dxftype() == "INSERT":
                            for v_att in getattr(v_ent, "attribs", []): texts.extend(_텍스트_데이터_추출(v_att))
                except Exception: pass
    except Exception: pass
    seen, out = set(), []
    for x, y, txt, h in texts:
        clean = _정리문자열(txt)
        key = (round(x, 2), round(y, 2), clean)
        if key not in seen:
            seen.add(key); out.append((float(x), float(y), clean, float(h)))
    return out

def _clean_title_only(title: str) -> str:
    clean = re.sub(r"\bA1\b|\bA3\b|\bA0\b|NONE|N/A|1\s?[/:,]\s?\d{1,4}", " ", title, flags=re.I)
    clean = re.sub(r"(?:축척|SCALE)?\s*\(\s*1\s*[:/]\s*\)", " ", clean, flags=re.I)
    clean = re.sub(r"(?:축척|SCALE)\s*1\s*[:/]", " ", clean, flags=re.I)
    clean = _clean_text_from_headers(clean)
    return clean

def _extract_scale_smart(cell_texts: List[Tuple[float, float, str, float]], header_a1_x: Optional[float] = None, header_a3_x: Optional[float] = None) -> Tuple[str, str]:
    a1_val, a3_val = "X", "X"
    scales, labels = [], {}
    for x, y, txt, h in cell_texts:
        u_txt = txt.upper()
        clean_txt = u_txt.replace(" ", "")
        m_a1 = re.search(r'A1.*?(1\s?[/:,]\s?\d{1,4}|NONE|N/A)', clean_txt)
        if m_a1 and a1_val == "X": a1_val = _축척_텍스트_정리(m_a1.group(1))
        m_a3 = re.search(r'A3.*?(1\s?[/:,]\s?\d{1,4}|NONE|N/A)', clean_txt)
        if m_a3 and a3_val == "X": a3_val = _축척_텍스트_정리(m_a3.group(1))
        
        if re.search(r'\bA1\b', u_txt): labels['A1'] = (x, y)
        if re.search(r'\bA3\b', u_txt): labels['A3'] = (x, y)
        for m in _축척_패턴.finditer(u_txt):
            val = _축척_텍스트_정리(m.group(0))
            if val != "X": scales.append((x, y, val))
            
    unique_scales, seen = [], set()
    for sx, sy, sval in scales:
        if (sx, sy, sval) not in seen:
            seen.add((sx, sy, sval)); unique_scales.append((sx, sy, sval))
            
    def dist(x1, y1, x2, y2): return math.sqrt((x1 - x2)**2 + (y1 - y2)**2)
    pairings = []
    for sx, sy, sval in unique_scales:
        d_a1 = dist(sx, sy, labels['A1'][0], labels['A1'][1]) if 'A1' in labels else (abs(sx - header_a1_x) if header_a1_x is not None else float('inf'))
        d_a3 = dist(sx, sy, labels['A3'][0], labels['A3'][1]) if 'A3' in labels else (abs(sx - header_a3_x) if header_a3_x is not None else float('inf'))
        if d_a1 != float('inf') or d_a3 != float('inf'):
            if d_a1 <= d_a3: pairings.append((d_a1, sval, 'A1'))
            else: pairings.append((d_a3, sval, 'A3'))
            
    pairings.sort(key=lambda p: p[0])
    for d, sval, target in pairings:
        if target == 'A1' and a1_val == "X": a1_val = sval
        elif target == 'A3' and a3_val == "X": a3_val = sval
        
    if unique_scales:
        unique_scales.sort(key=lambda item: item[0])
        if a1_val == "X" and a3_val == "X":
            if len(unique_scales) >= 2: a1_val, a3_val = unique_scales[0][2], unique_scales[1][2]
            else: a1_val = unique_scales[0][2]
        elif a1_val == "X" and a3_val != "X":
            for _, _, sval in unique_scales:
                if sval != a3_val: a1_val = sval; break
        elif a3_val == "X" and a1_val != "X":
            for _, _, sval in unique_scales:
                if sval != a1_val: a3_val = sval; break
    return a1_val, a3_val

# ============================================================================
# 2. 도면목록표 (DWG) 파싱 로직
# ============================================================================
def extract_dwg_list_table(dwg_path: str, block_name: str, roi_cfg: dict, base_w: float, base_h: float) -> pd.DataFrame:
    print(f"\n[LIST] DWG 도면목록표 분석 시작: {os.path.basename(dwg_path)}")
    데이터, 목표블록 = [], block_name.strip().lower()
    list_rois = roi_cfg.get('list_rois', [])
    if not list_rois: print("⚠️ [경고] 리습에서 지정된 목록표 단(ROI)이 없습니다. 전체 스캔을 시도합니다.")
    
    global_ignores_stripped = [h.replace(" ", "").upper() for h in GLOBAL_IGNORE_HEADERS] + ["A1", "A3", "A0"]
    
    try:
        doc = _cad_로드(Path(dwg_path))
        for layout in doc.layouts:
            도곽들 = [ins for ins in layout.query("INSERT") if 목표블록 in ins.dxf.name.lower()]
            if not 도곽들: continue
            모든텍스트 = _collect_layout_texts(layout)
            for 도곽 in 도곽들:
                ix, iy = float(도곽.dxf.insert.x), float(도곽.dxf.insert.y)
                xscale, yscale = abs(float(도곽.dxf.xscale)), abs(float(도곽.dxf.yscale))
                너비, 높이 = base_w * xscale, base_h * yscale
                rot_deg = getattr(도곽.dxf, 'rotation', 0.0)
                rad = math.radians(-rot_deg)
                cos_val, sin_val = math.cos(rad), math.sin(rad)

                target_ranges = list_rois if list_rois else [[0.0, 1.0, 0.0, 1.0]]
                
                # 리습에서 친 단(Box) 개수만큼 반복 스캔
                for roi_idx, roi in enumerate(target_ranges):
                    min_x, max_x = ix + (너비 * roi[0]), ix + (너비 * roi[1])
                    y_min, y_max = iy + (높이 * roi[2]), iy + (높이 * roi[3])
                    roi_w = max_x - min_x
                    
                    num_x_cands, title_x_cands, remark_x_cands = [], [], []
                    a1_matches, a3_matches = [], []
                    구역_텍스트 = []
                    
                    for t in 모든텍스트:
                        tx, ty, txt, th = t
                        dx, dy = tx - ix, ty - iy
                        unrot_x = ix + (dx * cos_val - dy * sin_val)
                        unrot_y = iy + (dx * sin_val + dy * cos_val)
                        
                        if min_x <= unrot_x <= max_x and y_min <= unrot_y <= y_max:
                            clean_t = txt.replace(" ", "").replace("\n", "").strip().upper()
                            
                            # 앵커(등대) 좌표 수집
                            if clean_t in ["도면번호", "도연번호", "DWG.NO", "DWG.NO.", "DWGNO", "DRAWINGNO", "번호"]: num_x_cands.append(unrot_x)
                            if clean_t in ["도면명", "DRAWINGTITLE", "TITLE", "도면명칭"]: title_x_cands.append(unrot_x)
                            if clean_t in ["비고", "REMARK", "REMARKS"]: remark_x_cands.append(unrot_x)
                            
                            if txt == "-" and th > roi_w * 0.8: continue
                            
                            if not _extract_drawing_number(txt):
                                if re.search(r"\bA1\b", txt.upper()): a1_matches.append((unrot_x, unrot_y))
                                if re.search(r"\bA3\b", txt.upper()): a3_matches.append((unrot_x, unrot_y))
                            
                            if any(ih == clean_t for ih in global_ignores_stripped): continue
                            구역_텍스트.append((unrot_x, unrot_y, txt, th))
                    
                    if not 구역_텍스트: continue
                    
                    header_num_x = sum(num_x_cands)/len(num_x_cands) if num_x_cands else min_x + (roi_w * 0.15)
                    header_title_x = sum(title_x_cands)/len(title_x_cands) if title_x_cands else min_x + (roi_w * 0.5)
                    header_remark_x = sum(remark_x_cands)/len(remark_x_cands) if remark_x_cands else max_x
                    
                    header_a1_x = sorted(a1_matches, key=lambda v: -v[1])[0][0] if a1_matches else None
                    header_a3_x = sorted(a3_matches, key=lambda v: -v[1])[0][0] if a3_matches else None
                            
                    tight_y_tol = 높이 * 0.012 
                    구역_텍스트.sort(key=lambda x: -x[1]) 
                    
                    sub_lines, curr_sub, curr_y = [], [], None
                    for t in 구역_텍스트:
                        if curr_y is None or abs(curr_y - t[1]) <= tight_y_tol:
                            curr_y = t[1]; curr_sub.append(t)
                        else:
                            curr_sub.sort(key=lambda x: x[0]); sub_lines.append({'y': curr_y, 'texts': curr_sub})
                            curr_y = t[1]; curr_sub = [t]
                    if curr_sub:
                        curr_sub.sort(key=lambda x: x[0]); sub_lines.append({'y': curr_y, 'texts': curr_sub})

                    rows, unassigned_sub_lines = [], []
                    
                    for sub in sub_lines:
                        full_str = " ".join([t[2] for t in sub['texts']])
                        
                        # [카테고리 스킵] "A0 [공통사항]" 등 버리기
                        is_category = False
                        if any(kw in full_str.replace(" ", "") for kw in CATEGORY_KEYWORDS): is_category = True
                        elif re.search(r"^[A-Z0-9\-_]*\s*[\[<【].+?[\]>】]\s*$", full_str): is_category = True
                        if is_category: continue

                        # [도면번호 추출]
                        raw_drw_no = _extract_drawing_number(full_str)
                        drw_no, raw_matched_str = "", ""

                        if raw_drw_no:
                            drw_no = _도면번호_세척(raw_drw_no)
                            raw_matched_str = raw_drw_no
                        else:
                            # 앵커 기반으로 좌측 텍스트 추정
                            num_texts = [t for t in sub['texts'] if abs(t[0] - header_num_x) <= abs(t[0] - header_title_x)]
                            if num_texts:
                                raw_left_str = " ".join([t[2] for t in num_texts])
                                fallback_match = re.sub(r"\s*[가-힣\[<【\(].*$", "", raw_left_str).strip("-_ ")
                                if not fallback_match: fallback_match = raw_left_str.strip()
                                if re.search(r"\d", fallback_match) and len(fallback_match) >= 3 and not re.search(r"[\[\]<>\(【】]", fallback_match):
                                    drw_no = _도면번호_세척(fallback_match)
                                    raw_matched_str = fallback_match

                        if drw_no:
                            rows.append({'anchor_y': sub['y'], 'sub_lines': [{'y': sub['y'], 'texts': sub['texts'], 'raw_drw_no': raw_matched_str}], 'drw_no': drw_no})
                        else:
                            unassigned_sub_lines.append({'y': sub['y'], 'texts': sub['texts']})

                    for sub in unassigned_sub_lines:
                        if not rows: continue
                        closest_row = min(rows, key=lambda r: abs(r['anchor_y'] - sub['y']))
                        if abs(closest_row['anchor_y'] - sub['y']) < 높이 * 0.04:
                            closest_row['sub_lines'].append(sub)

                    for row in rows:
                        row['sub_lines'].sort(key=lambda s: -s['y']) 
                        title_words, all_texts = [], []
                        
                        for sub in row['sub_lines']:
                            sub_texts_sorted = sorted(sub['texts'], key=lambda x: x[0])
                            
                            # [비고 차단막] 비고 앵커 주변 글자는 버림
                            title_texts = []
                            for t in sub_texts_sorted:
                                if header_remark_x and abs(t[0] - header_remark_x) < abs(t[0] - header_title_x): continue
                                title_texts.append(t)
                                
                            raw_left_str = " ".join([t[2] for t in title_texts])
                            
                            # [일련번호 절단기] 도면번호보다 왼쪽에 있는 글자(1, 2, 3 등)는 날림
                            title_overflow = raw_left_str
                            if sub.get('raw_drw_no') and sub['raw_drw_no'] in raw_left_str:
                                parts = raw_left_str.split(sub['raw_drw_no'], 1)
                                title_overflow = parts[1] if len(parts) > 1 else ""
                                
                            cleaned_line = _clean_title_only(title_overflow)
                            if cleaned_line: title_words.append(cleaned_line)
                            all_texts.extend(sub_texts_sorted)

                        번호 = row['drw_no']
                        명칭 = " ".join(title_words).strip()
                        current_dong = "공통"
                        extracted_dong = _extract_dong_from_title(명칭)
                        if extracted_dong:
                            current_dong = extracted_dong
                            임시_명칭 = 명칭.replace(current_dong, "")
                            임시_명칭 = re.sub(r"^[,\s]+|[,\s]+$", "", 임시_명칭).strip()
                            if 임시_명칭: 명칭 = 임시_명칭

                        a1, a3 = _extract_scale_smart(all_texts, header_a1_x, header_a3_x)
                        데이터.append({
                            "도면번호(LIST)": 번호, 
                            "구분_LIST(동)": current_dong if current_dong != "공통" else "", 
                            "도면명(LIST)": 명칭, 
                            "축척_A1(LIST)": a1, 
                            "축척_A3(LIST)": a3
                        })
    except Exception as e: print(f"[ERROR] 목록표 분석 중 오류: {e}")
    df = pd.DataFrame(데이터)
    if df.empty: return pd.DataFrame(columns=["도면번호(LIST)", "구분_LIST(동)", "도면명(LIST)", "축척_A1(LIST)", "축척_A3(LIST)"])
    return df.drop_duplicates(subset=["도면번호(LIST)"]).reset_index(drop=True)

# ============================================================================
# 3. 개별 도면 (DWG) 파싱
# ============================================================================
def _process_single_dwg(args: Tuple[str, str, dict, float, float]) -> Tuple[List[dict], str]:
    전체경로, 목표블록, roi_cfg, base_w, base_h = args
    파일명, 데이터, 에러메시지 = os.path.basename(전체경로), [], ""
    try:
        doc = _cad_로드(Path(전체경로))
        도곽_발견됨 = False

        for layout in doc.layouts:
            도곽들 = [ins for ins in layout.query("INSERT") if 목표블록 in ins.dxf.name.lower()]
            if not 도곽들: continue
            
            도곽_발견됨 = True
            모든텍스트 = _collect_layout_texts(layout)

            for 도곽 in 도곽들:
                ix, iy = float(도곽.dxf.insert.x), float(도곽.dxf.insert.y)
                xscale, yscale = abs(float(도곽.dxf.xscale)), abs(float(도곽.dxf.yscale))
                너비, 높이 = base_w * xscale, base_h * yscale
                rot_deg = getattr(도곽.dxf, 'rotation', 0.0)
                rad = math.radians(-rot_deg)
                cos_val, sin_val = math.cos(rad), math.sin(rad)

                def get_data_in_roi(roi):
                    x_min, x_max = ix + (너비 * roi[0]), ix + (너비 * roi[1])
                    y_min, y_max = iy + (높이 * roi[2]), iy + (높이 * roi[3])
                    
                    박스내글자 = []
                    for t in 모든텍스트:
                        tx, ty, txt, th = t
                        dx, dy = tx - ix, ty - iy
                        unrot_x = ix + (dx * cos_val - dy * sin_val)
                        unrot_y = iy + (dx * sin_val + dy * cos_val)
                        
                        if x_min <= unrot_x <= x_max and y_min <= unrot_y <= y_max:
                            if txt == "-" and th > (x_max - x_min) * 0.8: continue
                            박스내글자.append((unrot_x, unrot_y, txt, th)) 
                            
                    if not 박스내글자: return "", []
                    박스내글자.sort(key=lambda t: -t[1])
                    lines, current_line, current_y = [], [], None
                    y_tol = 높이 * 0.015

                    for t in 박스내글자:
                        if current_y is None: current_y = t[1]; current_line.append(t)
                        elif abs(current_y - t[1]) <= y_tol: current_line.append(t)
                        else:
                            current_line.sort(key=lambda x: x[0]); lines.append(" ".join([x[2] for x in current_line]))
                            current_y = t[1]; current_line = [t]
                    if current_line: current_line.sort(key=lambda x: x[0]); lines.append(" ".join([x[2] for x in current_line]))
                    return " ".join(lines), 박스내글자

                n_str, _ = get_data_in_roi(roi_cfg['num_roi'])
                t_str, _ = get_data_in_roi(roi_cfg['title_roi'])
                _, s_texts = get_data_in_roi(roi_cfg['scale_roi']) 

                n_str_clean = _clean_text_from_headers(n_str)
                t_str_clean = _clean_text_from_headers(t_str)

                번호_후보 = _extract_drawing_number(n_str_clean)
                if 번호_후보: 번호 = _도면번호_세척(번호_후보)
                else: 번호 = re.sub(r"\s*[가-힣\[<【\(].*$", "", n_str_clean).strip("-_ ")
                
                명칭 = t_str_clean
                if 번호 and 번호 in 명칭: 명칭 = 명칭.replace(번호, "")

                dwg_dong = _extract_dong_from_title(명칭)
                if dwg_dong:
                    임시_명칭 = 명칭.replace(dwg_dong, "")
                    임시_명칭 = re.sub(r"^[,\s]+|[,\s]+$", "", 임시_명칭).strip()
                    if 임시_명칭: 명칭 = 임시_명칭
                    
                명칭 = _clean_title_only(명칭)
                a1, a3 = _extract_scale_smart(s_texts)

                if 번호: 
                    데이터.append({
                        "파일명": 파일명, 
                        "도면번호(DWG)": 번호, 
                        "구분_DWG(동)": dwg_dong,
                        "도면명(DWG)": 명칭.strip(), 
                        "축척_A1(DWG)": a1, 
                        "축척_A3(DWG)": a3
                    })
        del doc
        if not 도곽_발견됨: return 데이터, "도곽 블록 없음"
    except Exception as e: 에러메시지 = str(e)
    return 데이터, 에러메시지

def extract_dwg_data_multiprocess(target_dirs: List[str], block_name: str, roi_cfg: dict, base_w: float, base_h: float) -> pd.DataFrame:
    모든_캐드파일 = []
    for d in target_dirs:
        폴더 = Path(d)
        if 폴더.exists(): 모든_캐드파일.extend([str(p) for p in 폴더.iterdir() if p.is_file() and p.suffix.lower() in [".dwg", ".dxf"]])
    캐드파일들 = sorted(list(set(모든_캐드파일)))
    if not 캐드파일들:
        print("[CAD ] 폴더 내에 처리할 도면 파일이 없습니다."); return pd.DataFrame(columns=["파일명", "도면번호(DWG)", "구분_DWG(동)", "도면명(DWG)", "축척_A1(DWG)", "축척_A3(DWG)"])

    print(f"\n[CAD ] 총 {len(캐드파일들)}개의 도면 분석 중... (터보 모드 가동 🚀)")
    최종_데이터 = []
    with concurrent.futures.ProcessPoolExecutor() as executor:
        futures = {executor.submit(_process_single_dwg, (path, block_name.strip().lower(), roi_cfg, base_w, base_h)): path for path in 캐드파일들}
        for i, future in enumerate(concurrent.futures.as_completed(futures), 1):
            경로 = futures[future]
            try:
                결과, 에러 = future.result()
                if 결과: 최종_데이터.extend(결과)
                print(f"   [{i}/{len(캐드파일들)}] {'완료' if 결과 else '패스'}: {os.path.basename(경로)} ({에러 if 에러 else '성공'})")
            except Exception as e:
                print(f"   [{i}/{len(캐드파일들)}] 시스템 오류: {os.path.basename(경로)} ({e})")
    
    if not 최종_데이터:
        return pd.DataFrame(columns=["파일명", "도면번호(DWG)", "구분_DWG(동)", "도면명(DWG)", "축척_A1(DWG)", "축척_A3(DWG)"])
    return pd.DataFrame(최종_데이터)

# ============================================================================
# 4. 리포트 생성
# ============================================================================
def build_report(list_df: pd.DataFrame, dwg_df: pd.DataFrame, out_path: str):
    if list_df.empty and dwg_df.empty:
        print("[알림] 추출된 데이터가 없어 엑셀 리포트를 생성하지 않습니다.")
        return

    lst, dwg = list_df.copy(), dwg_df.copy()
    
    if "도면번호(LIST)" not in lst.columns: lst["도면번호(LIST)"] = ""
    if "도면번호(DWG)" not in dwg.columns: dwg["도면번호(DWG)"] = ""
    if "구분_LIST(동)" not in lst.columns: lst["구분_LIST(동)"] = ""
    if "구분_DWG(동)" not in dwg.columns: dwg["구분_DWG(동)"] = ""

    lst["KEY"] = lst["도면번호(LIST)"].astype(str).str.upper().str.replace(r"[\s\-_]", "", regex=True)
    dwg["KEY"] = dwg["도면번호(DWG)"].astype(str).str.upper().str.replace(r"[\s\-_]", "", regex=True)
    
    결과 = pd.merge(lst, dwg, on="KEY", how="outer", indicator=True)
    결과["상태"] = 결과["_merge"].map({"both": "일치", "left_only": "DWG 누락", "right_only": "목록표 누락"})

    dong_mismatch_indices = set()
    for i in range(len(결과)):
        l_d = str(결과.at[i, "구분_LIST(동)"]).strip()
        d_d = str(결과.at[i, "구분_DWG(동)"]).strip()
        if l_d == "nan": l_d = ""
        if d_d == "nan": d_d = ""
        if l_d and d_d and l_d != d_d: dong_mismatch_indices.add(i + 2)

    prev_dong = ""
    dong_col_idx = 결과.columns.get_loc("구분_LIST(동)")
    for i in range(len(결과)):
        curr_dong = str(결과.iat[i, dong_col_idx]).strip()
        if curr_dong == "nan" or not curr_dong:
            prev_dong = ""; 결과.iat[i, dong_col_idx] = ""; continue
        if curr_dong == prev_dong: 결과.iat[i, dong_col_idx] = ""  
        else: prev_dong = curr_dong          

    prev_dwg_dong = ""
    dwg_dong_col_idx = 결과.columns.get_loc("구분_DWG(동)")
    for i in range(len(결과)):
        curr_dong = str(결과.iat[i, dwg_dong_col_idx]).strip()
        if curr_dong == "nan" or not curr_dong:
            prev_dwg_dong = ""; 결과.iat[i, dwg_dong_col_idx] = ""; continue
        if curr_dong == prev_dwg_dong: 결과.iat[i, dwg_dong_col_idx] = ""  
        else: prev_dwg_dong = curr_dong          

    cols = ["도면번호(LIST)", "구분_LIST(동)", "도면명(LIST)", "축척_A1(LIST)", "축척_A3(LIST)", 
            "도면번호(DWG)", "구분_DWG(동)", "도면명(DWG)", "축척_A1(DWG)", "축척_A3(DWG)", "파일명", "상태"]
    for c in cols: 
        if c not in 결과.columns: 결과[c] = ""
    
    결과[cols].fillna("X").to_excel(out_path, index=False)
    
    wb = load_workbook(out_path); ws = wb.active
    빨간색 = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")
    h = {cell.value: cell.column for cell in ws[1] if cell.value}
    
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, h["상태"]).value != "일치":
            for c in range(1, len(cols)+1): ws.cell(row, c).fill = 빨간색
        else:
            if row in dong_mismatch_indices:
                if h.get("구분_LIST(동)"): ws.cell(row, h.get("구분_LIST(동)")).fill = 빨간색
                if h.get("구분_DWG(동)"): ws.cell(row, h.get("구분_DWG(동)")).fill = 빨간색

            val_list = re.sub(r"[\s\-_]", "", str(ws.cell(row, h.get("도면번호(LIST)")).value).upper())
            val_dwg = re.sub(r"[\s\-_]", "", str(ws.cell(row, h.get("도면번호(DWG)")).value).upper())
            if val_list != val_dwg:
                ws.cell(row, h.get("도면번호(LIST)")).fill = 빨간색
                ws.cell(row, h.get("도면번호(DWG)")).fill = 빨간색
                
            name_list = str(ws.cell(row, h.get("도면명(LIST)")).value).replace(" ", "")
            name_dwg = str(ws.cell(row, h.get("도면명(DWG)")).value).replace(" ", "")
            if name_list != name_dwg:
                ws.cell(row, h.get("도면명(LIST)")).fill = 빨간색
                ws.cell(row, h.get("도면명(DWG)")).fill = 빨간색

            for s in ["A1", "A3"]:
                p_v = str(ws.cell(row, h[f"축척_{s}(LIST)"]).value).replace(" ","")
                d_v = str(ws.cell(row, h[f"축척_{s}(DWG)"]).value).replace(" ","")
                if p_v != d_v:
                    ws.cell(row, h[f"축척_{s}(LIST)"]).fill = 빨간색
                    ws.cell(row, h[f"축척_{s}(DWG)"]).fill = 빨간색

    wb.save(out_path)
    print(f"\n[XLSX] 리포트 저장 완료: {out_path}")

# ============================================================================
# 5. 메인 함수 
# ============================================================================
def main():
    print("=" * 72)
    print(" AutoDWG Cross-Checker v_6.8 (Kunwon Masterpiece - Multi-Zoning)")
    print("=" * 72)
    print(" Copyright (c) 2026 건원건축(Kunwon Architecture) & 김정현. All rights reserved.")
    print("=" * 72)

    check_oda_installation()

    blk_name = input("\n1. 도곽 블록 이름을 입력하세요: ").strip()
    roi_config = load_roi_config(blk_name)

    if not roi_config:
        print(f"\n[오류] '{blk_name}'에 대한 구역 설정 파일이 없습니다!")
        config_dir = os.path.join(os.environ.get('APPDATA', ''), 'AutoDWG_Checker')
        if os.path.exists(config_dir):
            saved_files = [f.replace('.json', '') for f in os.listdir(config_dir) if f.endswith('.json')]
            if saved_files:
                print("\n💡 [참고] 현재 저장되어 있는 도곽 목록은 다음과 같습니다:")
                for sf in saved_files: print(f"   - {sf}")
                print("\n(이름의 띄어쓰기, 대소문자, 괄호 모양이 정확히 일치해야 합니다!)")
        
        print("\n캐드에서 SET_ROI 명령어로 구역을 먼저 지정해 주세요.")
        input("\n엔터를 누르면 종료됩니다...")
        return

    base_w = float(roi_config.get('base_w', 841.0))
    base_h = float(roi_config.get('base_h', 594.0))

    print(f"\n[성공] '{blk_name}' 설정을 로드했습니다. (원본크기: {base_w}x{base_h})")

    print("\n2. 도면 목록표 DWG 파일의 경로를 입력하세요.")
    목록표_경로 = input("   경로: ").strip().strip('"')
    if not os.path.isfile(목록표_경로):
        print("[ERROR] 유효한 파일이 아닙니다. 종료합니다."); return

    dwg_dirs = []
    print("\n3. 검토할 개별 도면(DWG) 폴더 경로를 입력하세요. (여러 개 입력 가능)")
    print("   더 이상 없다면 'N'을 입력하세요.")
    while True:
        path_input = input(f"   [{len(dwg_dirs) + 1}번째 폴더] (또는 N): ").strip().strip('"')
        if path_input.upper() == 'N': break
        if os.path.isdir(path_input): dwg_dirs.append(path_input)

    print("-" * 72)

    if getattr(sys, 'frozen', False): 실행폴더 = os.path.dirname(sys.executable)
    else: 실행폴더 = os.path.dirname(os.path.abspath(__file__))
    최종_저장경로 = os.path.join(실행폴더, 리포트_이름)

    try:
        list_데이터 = extract_dwg_list_table(목록표_경로, blk_name, roi_config, base_w, base_h)
        dwg_데이터 = extract_dwg_data_multiprocess(dwg_dirs, blk_name, roi_config, base_w, base_h)

        build_report(list_데이터, dwg_데이터, 최종_저장경로)
        
        print("-" * 72)
        print(f"[DONE] 검토 완료! 리포트가 프로그램과 같은 폴더에 저장되었습니다.")
        
    except PermissionError:
        print("\n[ERROR] 엑셀 파일이 이미 켜져 있습니다. 창을 닫고 다시 실행해 주세요.")
    except Exception as e:
        print(f"\n[ERROR] 알 수 없는 오류 발생: {e}")

    print("=" * 72)
    input("\n[안내] 모든 작업이 끝났습니다. 엔터키를 누르면 창이 닫힙니다...")


if __name__ == "__main__":
    import multiprocessing
    multiprocessing.freeze_support()
    main()